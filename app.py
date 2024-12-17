import cv2
from ultralytics import YOLO
import openpyxl
from datetime import datetime
import streamlit as st
import tempfile
import os

# ------------------- INITIALIZATION -------------------

# Load the YOLO model
model_path = 'Freshnew100.pt'  # Path to your .pt model file
model = YOLO(model_path)  # Load the YOLO model

# Define label mapping (ensure it matches your YOLO training classes)
label_encoder = ['apple_fresh', 'apple_stale', 'onion_fresh', 'onion_stale',
                 'carrot_fresh', 'carrot_stale', 'tomato_fresh', 'tomato_stale']

# Define expected lifespan for each product
expected_life_span = {
    "apple": 7,    # Apple lasts 7 days when fresh
    "onion": 10,   # Onion lasts 10 days when fresh
    "carrot": 5,   # Carrot lasts 5 days when fresh
    "tomato": 3    # Tomato lasts 3 days when fresh
}

# Confidence threshold for filtering low-confidence predictions
CONFIDENCE_THRESHOLD = 0.5

# Excel file for storing fresh count
excel_file = "detection_fresh_count3.xlsx"

# Initialize or load Excel workbook and sheet
try:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["S No", "Product", "Fresh Count", "Last Detected Time", "Expected Life Span"])

# ------------------- HELPER FUNCTIONS -------------------

# Helper function to update fresh count in Excel
def update_fresh_count(product, is_fresh):
    lifespan = "N/A" if not is_fresh else expected_life_span.get(product, "Unknown")

    # Check if the product already exists in the sheet
    product_found = False
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[1].value == product:  # Check the "Product" column
            product_found = True
            if is_fresh:  # Only increment count if the product is fresh
                row[2].value += 1  # Increment the fresh count
            row[3].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Update the last detected time
            row[4].value = lifespan  # Update expected lifespan
            break

    if not product_found:
        # Add a new row for the product
        fresh_count = 1 if is_fresh else 0
        last_detected_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([sheet.max_row, product, fresh_count, last_detected_time, lifespan])

# Detect and classify products in an image
def detect_and_classify(image_path):
    image = cv2.imread(image_path)
    if image is None:
        st.error("Error: Image not found.")
        return None

    # Use the YOLO model for prediction
    results = model(image)

    predictions = []  # Store predictions to display on the frontend

    # Iterate through detected objects
    for result in results:
        boxes = result.boxes.xyxy  # Bounding box coordinates
        confidences = result.boxes.conf  # Confidence scores
        labels = result.boxes.cls  # Class indices

        for i, box in enumerate(boxes):
            confidence = confidences[i].item()  # Confidence score
            if confidence < CONFIDENCE_THRESHOLD:
                continue  # Skip low-confidence predictions

            label_idx = int(labels[i])  # Class index
            predicted_label = label_encoder[label_idx]  # Get label from mapping
            product, freshness = predicted_label.split('_')  # Split into product and freshness

            # Update fresh count in Excel
            is_fresh = (freshness == "fresh")
            update_fresh_count(product, is_fresh)

            # Add to predictions list
            predictions.append({
                "Product": product.capitalize(),
                "Freshness": freshness.capitalize(),
                "Confidence": f"{confidence:.2f}"
            })

    return predictions  # Return predictions list

# ------------------- STREAMLIT UI -------------------

def main():
    st.title("Fresh Product Detection with YOLOv8")
    st.write("Upload an image, and this app will detect fresh and stale products.")

    # File uploader
    uploaded_file = st.file_uploader("Choose an image...", type=["jpg", "jpeg", "png"])

    if uploaded_file is not None:
        # Save the uploaded image temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_image:
            temp_image.write(uploaded_file.read())
            image_path = temp_image.name

        # Detect and classify products
        st.write("Processing image...")
        predictions = detect_and_classify(image_path)

        if predictions:
            # Display predictions in a table
            st.write("### Detection Results")
            st.table(predictions)  # Streamlit table to display results

            # Save Excel file
            workbook.save(excel_file)
            st.success("Fresh count updated and saved.")

            # Add a download button for the updated Excel file
            with open(excel_file, "rb") as file:
                st.download_button(
                    label="Download Updated Excel Sheet",
                    data=file,
                    file_name="updated_fresh_count.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("No objects were detected in the image. Please try again.")

        # Cleanup: Delete temporary image
        os.remove(image_path)

if __name__ == "__main__":
    main()
